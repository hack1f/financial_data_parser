
import pandas as pd
import openpyxl
import numpy as np
import sqlite3
import re
from datetime import datetime
from decimal import Decimal
import locale


class ExcelProcessor:
    def __init__(self):
        self.workbooks = {}
        self.sheets = {}

    def load_files(self, file_paths):
        """Load multiple Excel files and store workbook information."""
        for path in file_paths:
            try:
                workbook = openpyxl.load_workbook(path)
                self.workbooks[path] = workbook
                self.sheets[path] = workbook.sheetnames
            except Exception as e:
                print(f"Error loading {path}: {str(e)}")

    def get_sheet_info(self):
        """Return information about loaded sheets."""
        info = {}
        for path, workbook in self.workbooks.items():
            info[path] = {
                'sheets': workbook.sheetnames,
                'dimensions': {
                    sheet: {
                        'rows': workbook[sheet].max_row,
                        'columns': workbook[sheet].max_column,
                        'column_names': [cell.value for cell in workbook[sheet][1]]
                    } for sheet in workbook.sheetnames
                }
            }
        return info

    def extract_data(self, file_path, sheet_name):
        """Extract data from specific sheet using pandas."""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df
        except Exception as e:
            print(f"Error extracting data from {sheet_name}: {str(e)}")
            return None

    def preview_data(self, file_path, sheet_name, rows=5):
        """Preview first n rows of the specified sheet."""
        df = self.extract_data(file_path, sheet_name)
        if df is not None:
            return df.head(rows)
        return None


class DataTypeDetector:
    def __init__(self):
        self.date_patterns = [
            r'^\d{1,2}/\d{1,2}/\d{4}$',  # MM/DD/YYYY or DD/MM/YYYY
            r'^\d{4}-\d{1,2}-\d{1,2}$',  # YYYY-MM-DD
            r'^Q[1-4]\s\d{4}$',  # Q1 2024
            r'^\w{3}-\d{2}$',  # Dec-23
            r'^\d+$'  # Excel serial date
        ]

        self.amount_patterns = [
            r'^\$?\s*-?\d{1,3}(,\d{3})*(\.\d{2})?$',  # $1,234.56 or -1,234.56
            r'^\€?\s*-?\d{1,3}(\.\d{3})*,\d{2}$',  # €1.234,56
            r'^\₹?\s*-?\d{1,3}(,\d{2})*(\.\d{2})?$',  # ₹1,23,456.78
            r'^\(?\d{1,3}(,\d{3})*\)?$',  # (1,234)
            r'^-?\d+\.?\d*[KMB]$'  # 1.23M, 2.5K
        ]

    def detect_column_type(self, column_data):
        """Detect data type of a column with confidence scores."""
        column_data = column_data.dropna()
        if len(column_data) == 0:
            return {'type': 'string', 'confidence': 1.0}

        date_score = self._check_date(column_data)
        number_score = self._check_number(column_data)
        string_score = 1.0 - max(date_score, number_score)

        if date_score >= max(number_score, string_score, 0.7):
            return {'type': 'date', 'confidence': date_score}
        elif number_score >= max(date_score, string_score, 0.7):
            return {'type': 'number', 'confidence': number_score}
        return {'type': 'string', 'confidence': string_score}

    def _check_date(self, column_data):
        """Check if column contains dates."""
        date_matches = 0
        total = len(column_data)

        for value in column_data:
            str_value = str(value)
            if any(re.match(pattern, str_value) for pattern in self.date_patterns):
                date_matches += 1
            elif isinstance(value, (datetime, pd.Timestamp)):
                date_matches += 1
            elif str_value.isdigit() and 40000 <= int(str_value) <= 50000:  # Excel serial date range
                date_matches += 1

        return date_matches / total if total > 0 else 0.0

    def _check_number(self, column_data):
        """Check if column contains numbers."""
        number_matches = 0
        total = len(column_data)

        for value in column_data:
            str_value = str(value).replace(' ', '')
            if any(re.match(pattern, str_value) for pattern in self.amount_patterns):
                number_matches += 1
            elif isinstance(value, (int, float, Decimal)):
                number_matches += 1

        return number_matches / total if total > 0 else 0.0


class FormatParser:
    def __init__(self):
        locale.setlocale(locale.LC_ALL, '')  # Use locale module directly

    def parse_amount(self, value):
        """Parse various amount formats to Decimal."""
        if pd.isna(value):
            return None

        str_value = str(value).strip()

        try:
            # Handle K/M/B abbreviations
            if str_value[-1] in ['K', 'M', 'B']:
                multiplier = {'K': 1000, 'M': 1000000, 'B': 1000000000}
                return Decimal(str_value[:-1]) * multiplier[str_value[-1]]

            # Handle parentheses for negative
            if str_value.startswith('(') and str_value.endswith(')'):
                str_value = '-' + str_value[1:-1]

            # Remove currency symbols and commas
            str_value = str_value.replace('$', '').replace('€', '').replace('₹', '')
            str_value = str_value.replace(',', '').replace(' ', '')

            return Decimal(str_value)
        except:
            return None

    def parse_date(self, value):
        """Parse various date formats to datetime."""
        if pd.isna(value):
            return None

        str_value = str(value).strip()

        try:
            # Handle Excel serial dates
            if str_value.isdigit() and 40000 <= int(str_value) <= 50000:
                return pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(str_value))

            # Handle quarter format
            if re.match(r'^Q[1-4]\s\d{4}$', str_value):
                year = int(str_value[-4:])
                quarter = int(str_value[1])
                month = {1: '01', 2: '04', 3: '07', 4: '10'}
                return pd.to_datetime(f'{year}-{month[quarter]}-01')

            # Try standard date parsing
            return pd.to_datetime(str_value)
        except:
            return None


class FinancialDataStore:
    def __init__(self):
        self.data = {}
        self.indexes = {}
        self.metadata = {}
        self.conn = sqlite3.connect(':memory:')

    def add_dataset(self, name, df, column_types):
        """Store dataset with indexes and metadata."""
        self.data[name] = df
        self.metadata[name] = column_types

        # Create SQLite table
        df.to_sql(name, self.conn, if_exists='replace', index=False)

        # Create indexes
        self.indexes[name] = {
            'date_index': {},
            'amount_index': {},
            'category_index': {}
        }

        for col in df.columns:
            if column_types.get(col, {}).get('type') == 'date':
                self.indexes[name]['date_index'][col] = df[col].to_dict()
            elif column_types.get(col, {}).get('type') == 'number':
                self.indexes[name]['amount_index'][col] = df[col].to_dict()
            else:
                self.indexes[name]['category_index'][col] = df[col].to_dict()

    def query_by_criteria(self, dataset_name, filters):
        """Query data with multiple criteria."""
        query = f"SELECT * FROM {dataset_name} WHERE "
        conditions = []
        params = []

        for column, condition in filters.items():
            if isinstance(condition, tuple):  # Range query
                conditions.append(f"{column} BETWEEN ? AND ?")
                params.extend(condition)
            else:
                conditions.append(f"{column} = ?")
                params.append(condition)

        query += " AND ".join(conditions)
        return pd.read_sql_query(query, self.conn, params=params)

    def aggregate_data(self, dataset_name, group_by, measures):
        """Perform aggregation on the dataset."""
        query = f"SELECT {group_by}, "
        agg_funcs = []

        for measure, agg in measures.items():
            agg_funcs.append(f"{agg}({measure}) as {measure}_{agg}")

        query += ", ".join(agg_funcs)
        query += f" FROM {dataset_name} GROUP BY {group_by}"

        return pd.read_sql_query(query, self.conn)


# Example usage and testing
if __name__ == "__main__":
    # Initialize components
    processor = ExcelProcessor()
    detector = DataTypeDetector()
    parser = FormatParser()
    store = FinancialDataStore()

    # Test file processing
    files = [r"C:\Users\PMLS\Desktop\sample\KH_Bank.XLSX",r"C:\Users\PMLS\Desktop\sample\Customer_Ledger_Entries_FULL.xlsx"]
    processor.load_files(files)

    # Get sheet info
    info = processor.get_sheet_info()
    for file, details in info.items():
        print(f"\nFile: {file}")
        for sheet, dims in details['dimensions'].items():
            print(f"Sheet: {sheet}, Rows: {dims['rows']}, Columns: {dims['columns']}")
            print(f"Column Names: {dims['column_names']}")

    # Test data type detection and parsing
    for file in files:
        for sheet in processor.sheets.get(file, []):
            df = processor.extract_data(file, sheet)
            if df is not None:
                column_types = {}
                for col in df.columns:
                    column_types[col] = detector.detect_column_type(df[col])

                    # Parse numbers and dates
                    if column_types[col]['type'] == 'number':
                        df[col] = df[col].apply(parser.parse_amount)
                    elif column_types[col]['type'] == 'date':
                        df[col] = df[col].apply(parser.parse_date)

                # Store data
                store.add_dataset(f"{file}_{sheet}", df, column_types)

                # Test queries
                results = store.query_by_criteria(
                    f"{file}_{sheet}",
                    {'amount': (1000, 10000)}  # Example range query
                )
                print(f"\nQuery Results for {file}_{sheet}:")
                print(results.head())

                # Test aggregation
                agg_results = store.aggregate_data(
                    f"{file}_{sheet}",
                    group_by='category',
                    measures={'amount': 'SUM'}
                )
                print(f"\nAggregation Results for {file}_{sheet}:")
                print(agg_results.head())
